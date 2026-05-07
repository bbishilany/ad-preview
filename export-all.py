#!/usr/bin/env python3
"""
Export ALL ad-preview data (data.json + Supabase feedback/images/review_requests)
into a single Excel workbook for backup.

Usage: python export-all.py
Output: ~/Desktop/Confluence FP - Ad Preview Backup.xlsx
"""

import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from supabase import create_client
except ImportError:
    print("Error: supabase-py not installed. Run: pip install supabase")
    sys.exit(1)


# ── Paths ────────────────────────────────────────────────────────────────────
PROJECT_DIR = Path(__file__).parent
DATA_JSON = PROJECT_DIR / "confluence-fp" / "data.json"
OUTPUT_PATH = Path.home() / "Desktop" / "Confluence FP - Ad Preview Backup.xlsx"

# ── Styles ───────────────────────────────────────────────────────────────────
NAVY = "152D4B"
GOLD = "C9A84C"
LIGHT_BG = "F5F5F5"
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
gold_font = Font(name="Arial", bold=True, color=NAVY, size=12)
gold_fill = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def style_header_row(ws, cols):
    for col_idx in range(1, cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border


def style_data_rows(ws, start_row, end_row, cols):
    for row in range(start_row, end_row + 1):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.alignment = wrap_align
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)


def auto_width(ws, cols, max_width=60):
    for col_idx in range(1, cols + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    longest = max(len(l) for l in lines)
                    max_len = max(max_len, longest)
        width = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)


def load_env():
    for env_path in [
        Path.home() / "Projects" / "mileage" / ".env",
        Path.home() / "Projects" / "mileage-os" / ".env",
    ]:
        if env_path.exists():
            for line in env_path.read_text().splitlines():
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, _, val = line.partition("=")
                    os.environ.setdefault(key.strip(), val.strip())


def get_supabase():
    load_env()
    url = os.environ.get("SUPABASE_URL", "")
    key = os.environ.get("SUPABASE_SERVICE_KEY", "")
    if not url or not key:
        print("Warning: SUPABASE_URL / SUPABASE_SERVICE_KEY not set. Skipping Supabase tables.")
        return None
    return create_client(url, key)


def write_campaign_sheet(wb, data):
    """Sheet 1: Campaign metadata."""
    ws = wb.active
    ws.title = "Campaign"
    rows = [
        ("Field", "Value"),
        ("Campaign", data.get("campaign", "")),
        ("Event Date", data.get("event_date", "")),
        ("Venue", data.get("venue", "")),
        ("Co-Presenters", data.get("co_presenters", "")),
        ("Total Ads", str(len(data.get("ads", [])))),
        ("Export Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for r, (field, value) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=field)
        ws.cell(row=r, column=2, value=value)
    style_header_row(ws, 2)
    style_data_rows(ws, 2, len(rows), 2)
    auto_width(ws, 2)


def write_ad_copy_sheet(wb, ads):
    """Sheet 2: All ad copy in a flat table."""
    ws = wb.create_sheet("Ad Copy")
    headers = [
        "Ad #", "Week", "Title", "Type", "Format", "Placements",
        "Primary Text", "Headline", "Description", "CTA",
        "Hook", "Alt Headlines", "Arc"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, ad in enumerate(ads, 2):
        ws.cell(row=row_idx, column=1, value=ad.get("num"))
        ws.cell(row=row_idx, column=2, value=ad.get("week"))
        ws.cell(row=row_idx, column=3, value=ad.get("title"))
        ws.cell(row=row_idx, column=4, value=ad.get("type"))
        ws.cell(row=row_idx, column=5, value=ad.get("format"))
        ws.cell(row=row_idx, column=6, value=", ".join(ad.get("placements", [])))
        ws.cell(row=row_idx, column=7, value=ad.get("primary_text"))
        ws.cell(row=row_idx, column=8, value=ad.get("headline"))
        ws.cell(row=row_idx, column=9, value=ad.get("description"))
        ws.cell(row=row_idx, column=10, value=ad.get("cta"))
        ws.cell(row=row_idx, column=11, value=ad.get("hook"))
        ws.cell(row=row_idx, column=12, value="\n".join(ad.get("alt_headlines", [])))
        ws.cell(row=row_idx, column=13, value=ad.get("arc"))

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, len(ads) + 1, cols)
    auto_width(ws, cols)


def write_carousel_sheet(wb, ads):
    """Sheet 3: Carousel card copy."""
    ws = wb.create_sheet("Carousel Cards")
    headers = ["Ad #", "Card #", "Label", "Caption"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row_idx = 2
    for ad in ads:
        for card in ad.get("cards", []):
            ws.cell(row=row_idx, column=1, value=ad.get("num"))
            ws.cell(row=row_idx, column=2, value=card.get("num"))
            ws.cell(row=row_idx, column=3, value=card.get("label"))
            ws.cell(row=row_idx, column=4, value=card.get("caption"))
            row_idx += 1

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, row_idx - 1, cols)
    auto_width(ws, cols)


def write_stories_sheet(wb, ads):
    """Sheet 4: Stories metadata."""
    ws = wb.create_sheet("Stories")
    headers = ["Ad #", "Image Direction", "Text Overlay", "Secondary Text", "CTA Sticker"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    row_idx = 2
    for ad in ads:
        stories = ad.get("stories")
        if stories:
            ws.cell(row=row_idx, column=1, value=ad.get("num"))
            ws.cell(row=row_idx, column=2, value=stories.get("image", ""))
            ws.cell(row=row_idx, column=3, value=stories.get("text_overlay", ""))
            ws.cell(row=row_idx, column=4, value=stories.get("secondary_text", ""))
            ws.cell(row=row_idx, column=5, value=stories.get("cta_sticker", ""))
            row_idx += 1

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, row_idx - 1, cols)
    auto_width(ws, cols)


def write_images_sheet(wb, ads):
    """Sheet 5: Image filenames and creative direction."""
    ws = wb.create_sheet("Images & Direction")
    headers = ["Ad #", "Title", "Image Files", "Image Creative Direction"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    for row_idx, ad in enumerate(ads, 2):
        ws.cell(row=row_idx, column=1, value=ad.get("num"))
        ws.cell(row=row_idx, column=2, value=ad.get("title"))
        ws.cell(row=row_idx, column=3, value="\n".join(ad.get("images", [])))
        ws.cell(row=row_idx, column=4, value=ad.get("image_direction", ""))

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, len(ads) + 1, cols)
    auto_width(ws, cols)


def write_feedback_sheet(wb, sb):
    """Sheet 6: All feedback from Supabase."""
    ws = wb.create_sheet("Feedback")
    headers = [
        "ID", "Ad #", "Campaign", "Reviewer Name", "Reviewer Role",
        "Verdict", "Comment", "Created At", "Pulled At"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    if not sb:
        ws.cell(row=2, column=1, value="(Supabase not connected)")
        style_header_row(ws, len(headers))
        auto_width(ws, len(headers))
        return 0

    result = (
        sb.table("feedback")
        .select("*")
        .eq("campaign_id", "confluence-fp")
        .order("ad_num")
        .order("created_at")
        .execute()
    )
    rows = result.data or []
    print(f"  Feedback rows: {len(rows)}")

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("id"))
        ws.cell(row=row_idx, column=2, value=row.get("ad_num"))
        ws.cell(row=row_idx, column=3, value=row.get("campaign_id"))
        ws.cell(row=row_idx, column=4, value=row.get("reviewer_name", ""))
        ws.cell(row=row_idx, column=5, value=row.get("reviewer_role", ""))
        ws.cell(row=row_idx, column=6, value=row.get("verdict", ""))
        ws.cell(row=row_idx, column=7, value=row.get("comment", ""))
        ws.cell(row=row_idx, column=8, value=row.get("created_at", ""))
        ws.cell(row=row_idx, column=9, value=row.get("pulled_at", ""))

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, len(rows) + 1, cols)
    auto_width(ws, cols)
    return len(rows)


def write_image_overrides_sheet(wb, sb):
    """Sheet 7: Image overrides from Supabase."""
    ws = wb.create_sheet("Image Overrides")
    headers = ["ID", "Campaign", "Ad #", "Slot", "Public URL", "Storage Path", "Created At"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    if not sb:
        ws.cell(row=2, column=1, value="(Supabase not connected)")
        style_header_row(ws, len(headers))
        auto_width(ws, len(headers))
        return 0

    result = (
        sb.table("ad_images")
        .select("*")
        .eq("campaign_id", "confluence-fp")
        .order("ad_num")
        .execute()
    )
    rows = result.data or []
    print(f"  Image override rows: {len(rows)}")

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("id"))
        ws.cell(row=row_idx, column=2, value=row.get("campaign_id"))
        ws.cell(row=row_idx, column=3, value=row.get("ad_num"))
        ws.cell(row=row_idx, column=4, value=row.get("slot"))
        ws.cell(row=row_idx, column=5, value=row.get("public_url", ""))
        ws.cell(row=row_idx, column=6, value=row.get("storage_path", ""))
        ws.cell(row=row_idx, column=7, value=row.get("created_at", ""))

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, len(rows) + 1, cols)
    auto_width(ws, cols)
    return len(rows)


def write_review_requests_sheet(wb, sb):
    """Sheet 8: Review board requests from Supabase."""
    ws = wb.create_sheet("Review Requests")
    headers = ["ID", "Campaign", "Requested By", "Email", "Message", "Status", "Created At"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    if not sb:
        ws.cell(row=2, column=1, value="(Supabase not connected)")
        style_header_row(ws, len(headers))
        auto_width(ws, len(headers))
        return 0

    try:
        result = (
            sb.table("review_requests")
            .select("*")
            .eq("campaign_id", "confluence-fp")
            .order("created_at")
            .execute()
        )
        rows = result.data or []
    except Exception as e:
        print(f"  (table not found or empty: {e})")
        rows = []
    print(f"  Review request rows: {len(rows)}")

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row.get("id"))
        ws.cell(row=row_idx, column=2, value=row.get("campaign_id"))
        ws.cell(row=row_idx, column=3, value=row.get("requested_by", ""))
        ws.cell(row=row_idx, column=4, value=row.get("email", ""))
        ws.cell(row=row_idx, column=5, value=row.get("message", ""))
        ws.cell(row=row_idx, column=6, value=row.get("status", ""))
        ws.cell(row=row_idx, column=7, value=row.get("created_at", ""))

    cols = len(headers)
    style_header_row(ws, cols)
    style_data_rows(ws, 2, len(rows) + 1, cols)
    auto_width(ws, cols)
    return len(rows)


def main():
    # Load data.json
    if not DATA_JSON.exists():
        print(f"Error: {DATA_JSON} not found")
        sys.exit(1)

    data = json.loads(DATA_JSON.read_text(encoding="utf-8"))
    ads = data.get("ads", [])
    print(f"Loaded {len(ads)} ads from data.json")

    # Connect Supabase
    sb = get_supabase()
    if sb:
        print("Connected to Supabase")

    # Build workbook
    wb = Workbook()

    print("\nWriting sheets...")
    print("  1. Campaign metadata")
    write_campaign_sheet(wb, data)

    print("  2. Ad Copy")
    write_ad_copy_sheet(wb, ads)

    print("  3. Carousel Cards")
    write_carousel_sheet(wb, ads)

    print("  4. Stories")
    write_stories_sheet(wb, ads)

    print("  5. Images & Direction")
    write_images_sheet(wb, ads)

    print("  6. Feedback (Supabase)")
    fb_count = write_feedback_sheet(wb, sb)

    print("  7. Image Overrides (Supabase)")
    img_count = write_image_overrides_sheet(wb, sb)

    print("  8. Review Requests (Supabase)")
    rr_count = write_review_requests_sheet(wb, sb)

    # Save
    wb.save(str(OUTPUT_PATH))
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"\nSummary:")
    print(f"  Ads: {len(ads)}")
    print(f"  Feedback entries: {fb_count}")
    print(f"  Image overrides: {img_count}")
    print(f"  Review requests: {rr_count}")


if __name__ == "__main__":
    main()
